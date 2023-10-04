import numpy as np
import openpyxl
from openpyxl import Workbook

def generate_fibonacci(n):
    fib_array = np.zeros(n, dtype=int)
    fib_array[0] = 0
    if n > 1:
        fib_array[1] = 1
        for i in range(2, n):
            fib_array[i] = fib_array[i-1] + fib_array[i-2]
    return fib_array

def write_to_excel(data_array, filename, sheetname=None, start_cell='A1'):
    try:
        # Charge le classeur s'il existe, sinon il en crée un nouveau
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            wb = Workbook()

        # Crée une nouvelle feuille si aucune n'est mentionné, sinon on utilise la feuille mentionné
        # ou en crée une nouvelle avec le nom donné
        if sheetname is None:
            ws = wb.create_sheet()
        else:
            ws = wb[sheetname] if sheetname in wb.sheetnames else wb.create_sheet(sheetname)
        
        # Obtient la rangée et la colonne de départ
        start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
        start_col = openpyxl.utils.column_index_from_string(start_col)

        # Vérifie si le tableau est à 1-D ou 2D et écrit les valeurs
        if data_array.ndim == 1:
            for i, item in enumerate(data_array):
                ws.cell(row=start_row + i, column=start_col, value=item)
        elif data_array.ndim == 2:
            for i in range(data_array.shape[0]):
                for j in range(data_array.shape[1]):
                    ws.cell(row=start_row + i, column=start_col + j, value=data_array[i, j])
        else:
            raise ValueError("Data array must be 1D or 2D.")

        # Sauvegarde le classeur
        wb.save(filename)
        print(f"Data written to {filename} successfully!")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    n = 10
    fibonacci_numbers = generate_fibonacci(n)

    # Writing 1D array to Excel
    write_to_excel(fibonacci_numbers, 'example.xlsx', 'Fibonacci_1D', 'A1')

    # Reshaping to 2D and writing to Excel
    fib_matrix = fibonacci_numbers.reshape((5, 2))
    write_to_excel(fib_matrix, 'example.xlsx', 'Fibonacci_2D', 'A1')
